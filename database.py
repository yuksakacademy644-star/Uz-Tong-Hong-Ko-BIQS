import sqlite3
import datetime
from config import DB_PATH

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    cursor = conn.cursor()

    # Users Table
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS users (
        telegram_id INTEGER PRIMARY KEY,
        full_name TEXT,
        username TEXT,
        phone TEXT,
        language TEXT DEFAULT 'ru',
        invite_code TEXT,
        shop_name TEXT,
        master_name TEXT,
        registered_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        role TEXT DEFAULT 'worker',
        permissions TEXT DEFAULT ''
    )
    """)

    # Invite Codes Table
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS invite_codes (
        code TEXT PRIMARY KEY,
        shop_name TEXT NOT NULL,
        master_name TEXT NOT NULL,
        created_by INTEGER,
        used_count INTEGER DEFAULT 0,
        max_uses INTEGER DEFAULT 9999,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        target_role TEXT DEFAULT 'worker'
    )
    """)

    # Test Results Table
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS test_results (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_telegram_id INTEGER,
        score INTEGER,
        total_questions INTEGER,
        percentage REAL,
        time_taken_seconds INTEGER,
        mistakes TEXT,
        completed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (user_telegram_id) REFERENCES users (telegram_id)
    )
    """)

    # Attacks Table
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS attacks (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        telegram_id INTEGER,
        attempt_details TEXT,
        attack_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """)

    # Migrations for existing DB
    cursor.execute("PRAGMA table_info(users)")
    cols = [r[1] for r in cursor.fetchall()]
    if 'role' not in cols:
        cursor.execute("ALTER TABLE users ADD COLUMN role TEXT DEFAULT 'worker'")
    if 'permissions' not in cols:
        cursor.execute("ALTER TABLE users ADD COLUMN permissions TEXT DEFAULT ''")

    cursor.execute("PRAGMA table_info(invite_codes)")
    code_cols = [r[1] for r in cursor.fetchall()]
    if 'target_role' not in code_cols:
        cursor.execute("ALTER TABLE invite_codes ADD COLUMN target_role TEXT DEFAULT 'worker'")

    conn.commit()
    conn.close()
    cleanup_old_test_results()
    init_banned_prompts_table()

def init_banned_prompts_table():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS banned_prompts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        pattern TEXT UNIQUE NOT NULL,
        category TEXT DEFAULT 'general',
        added_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """)
    conn.commit()

    cursor.execute("SELECT COUNT(*) FROM banned_prompts")
    if cursor.fetchone()[0] == 0:
        default_patterns = [
            # Profanity / Abusive terms
            ("блять", "profanity"), ("сука", "profanity"), ("хуй", "profanity"),
            ("пизд", "profanity"), ("нах", "profanity"), ("долбоеб", "profanity"),
            ("гандон", "profanity"), ("пидор", "profanity"), ("dalbayob", "profanity"),
            ("jallap", "profanity"), ("skay", "profanity"), ("qotoq", "profanity"),
            ("sikay", "profanity"), ("amcha", "profanity"),
            # Prompt injection / Attack strings
            ("ignore previous instructions", "prompt_injection"),
            ("forget all previous instructions", "prompt_injection"),
            ("you are now dan", "prompt_injection"),
            ("jailbreak", "prompt_injection"),
            ("select * from", "sql_injection"),
            ("drop table", "sql_injection"),
            ("union select", "sql_injection"),
            ("<script>", "xss"),
            ("eval(", "code_injection"),
            ("exec(", "code_injection"),
            ("system(", "code_injection"),
            ("rm -rf", "command_injection"),
            # Links
            ("http://", "link"), ("https://", "link"), ("t.me/", "link"),
            ("tg://", "link"), ("bit.ly", "link"), ("tinyurl.com", "link"),
            ("www.", "link"),
            # Honeypot / Unauthorized Probing
            ("/admin", "honeypot"), ("админ", "honeypot"), ("/panel", "honeypot"),
            ("панель", "honeypot"), ("/setup", "honeypot"), ("secret_admin_panel", "honeypot"),
            ("bot_token", "honeypot"),
            # Obfuscation & Malicious Payloads
            ("base64.b64decode", "code_injection"),
            ("aHR0cHM6Ly9tYWxpY2lvdXM", "code_injection"),
            ("payload =", "code_injection"),
            ("requests.post", "code_injection"),
            ("limit_req_zone", "attack_pattern"),
            ("binary_remote_addr", "attack_pattern")
        ]
        cursor.executemany("INSERT OR IGNORE INTO banned_prompts (pattern, category) VALUES (?, ?)", default_patterns)
        conn.commit()

    # Always ensure new patterns are inserted even if table already exists
    extra_patterns = [
        ("/admin", "honeypot"), ("админ", "honeypot"), ("/panel", "honeypot"),
        ("панель", "honeypot"), ("/setup", "honeypot"), ("secret_admin_panel", "honeypot"),
        ("bot_token", "honeypot"), ("base64.b64decode", "code_injection"),
        ("aHR0cHM6Ly9tYWxpY2lvdXM", "code_injection"), ("payload =", "code_injection"),
        ("requests.post", "code_injection"), ("limit_req_zone", "attack_pattern"),
        ("binary_remote_addr", "attack_pattern")
    ]
    cursor.executemany("INSERT OR IGNORE INTO banned_prompts (pattern, category) VALUES (?, ?)", extra_patterns)
    conn.commit()
    conn.close()



def cleanup_old_test_results():
    """
    Test results are permanently preserved for monthly reporting & historical archives.
    Automatic deletion is disabled per user requirements.
    """
    pass


def clear_all_invite_codes():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM invite_codes")
    conn.commit()
    conn.close()

# User DB functions
def get_user(telegram_id: int):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM users WHERE telegram_id = ?", (telegram_id,))
    row = cursor.fetchone()
    conn.close()
    return dict(row) if row else None

def create_user(telegram_id: int, full_name: str, username: str, phone: str, invite_code: str, shop_name: str, language: str = 'ru', role: str = 'worker'):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
    INSERT OR REPLACE INTO users (telegram_id, full_name, username, phone, language, invite_code, shop_name, master_name, role)
    VALUES (?, ?, ?, ?, ?, ?, ?, '', ?)
    """, (telegram_id, full_name, username, phone, language, invite_code, shop_name, role))
    
    # Update code used count
    cursor.execute("UPDATE invite_codes SET used_count = used_count + 1 WHERE code = ?", (invite_code,))
    
    conn.commit()
    conn.close()

def update_user_language(telegram_id: int, language: str):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE users SET language = ? WHERE telegram_id = ?", (language, telegram_id))
    conn.commit()
    conn.close()

def delete_user(telegram_id: int) -> bool:
    """Delete a user and all their test results. Returns True if user was found and deleted."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT telegram_id FROM users WHERE telegram_id = ?", (telegram_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return False
    cursor.execute("DELETE FROM test_results WHERE user_telegram_id = ?", (telegram_id,))
    cursor.execute("DELETE FROM attacks WHERE telegram_id = ?", (telegram_id,))
    cursor.execute("DELETE FROM users WHERE telegram_id = ?", (telegram_id,))
    conn.commit()
    conn.close()
    return True

# Invite Code DB functions
def get_invite_code(code: str):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM invite_codes WHERE code = ?", (code.strip().upper(),))
    row = cursor.fetchone()
    conn.close()
    return dict(row) if row else None

def add_invite_code(code: str, shop_name: str, master_name: str, created_by: int, target_role: str = 'worker'):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
    INSERT OR REPLACE INTO invite_codes (code, shop_name, master_name, created_by, target_role)
    VALUES (?, ?, ?, ?, ?)
    """, (code.strip().upper(), shop_name, master_name, created_by, target_role))
    conn.commit()
    conn.close()

def get_all_invite_codes():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM invite_codes ORDER BY created_at DESC")
    rows = cursor.fetchall()
    conn.close()
    return [dict(r) for r in rows]

# ─────────────────────────────────────────────────────────────────
# Role Hierarchy:
#   superadmin > admin > nachalnik > master > brigadir > worker
# ─────────────────────────────────────────────────────────────────
ROLE_RANK = {
    'superadmin': 100,
    'admin':      90,
    'director':   80,
    'quality':    75,
    'engineer':   75,
    'nachalnik':  70,
    'master':     50,
    'brigadir':   30,
    'worker':     10,
}
MANAGEMENT_ROLES = ('nachalnik', 'master', 'brigadir', 'quality', 'director', 'engineer')

def get_shop_workers(shop_name: str):
    """Returns all workers in a shop (for nachalnik — full shop, excluding admins)."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
    SELECT u.telegram_id, u.full_name, u.username, u.phone, u.shop_name, u.master_name,
           u.role, u.invite_code,
           COALESCE(MAX(tr.percentage), 0) as best_score,
           COUNT(tr.id) as tests_completed,
           (SELECT mistakes FROM test_results tr2
            WHERE tr2.user_telegram_id = u.telegram_id
            ORDER BY completed_at DESC LIMIT 1) as latest_mistakes
    FROM users u
    LEFT JOIN test_results tr ON u.telegram_id = tr.user_telegram_id
    WHERE u.shop_name = ? AND u.role NOT IN ('superadmin','admin')
    GROUP BY u.telegram_id
    ORDER BY
        CASE u.role
            WHEN 'nachalnik' THEN 1
            WHEN 'master'    THEN 2
            WHEN 'brigadir'  THEN 3
            ELSE 4
        END,
        best_score DESC
    """, (shop_name,))
    rows = cursor.fetchall()
    conn.close()
    return [dict(r) for r in rows]

def get_subordinates(user_id: int):
    """
    Returns the team visible to this user based on role:
    - nachalnik  → everyone in same shop (except admin/superadmin)
    - master     → brigadirs + workers in same shop where master_name matches
    - brigadir   → workers in same shop where master_name matches
    - admin/superadmin → all workers
    """
    import config as _cfg
    user = get_user(user_id)
    if not user:
        return []
    role = user.get('role', 'worker')
    shop = user.get('shop_name', '')
    name = user.get('full_name', '')

    conn = get_db()
    cursor = conn.cursor()

    base_select = """
    SELECT u.telegram_id, u.full_name, u.username, u.phone, u.shop_name, u.master_name,
           u.role, u.invite_code,
           COALESCE(MAX(tr.percentage), 0) as best_score,
           COUNT(tr.id) as tests_completed,
           (SELECT mistakes FROM test_results tr2
            WHERE tr2.user_telegram_id = u.telegram_id
            ORDER BY completed_at DESC LIMIT 1) as latest_mistakes
    FROM users u
    LEFT JOIN test_results tr ON u.telegram_id = tr.user_telegram_id
    """
    order = """
    GROUP BY u.telegram_id
    ORDER BY
        CASE u.role
            WHEN 'nachalnik' THEN 1
            WHEN 'master'    THEN 2
            WHEN 'brigadir'  THEN 3
            ELSE 4
        END,
        best_score DESC
    """

    if user_id in _cfg.ADMIN_IDS or role in ('superadmin', 'admin', 'director', 'quality', 'engineer'):
        cursor.execute(base_select + "WHERE u.role NOT IN ('superadmin','admin')" + order)
    elif role == 'nachalnik':
        cursor.execute(base_select + "WHERE u.shop_name = ? AND u.role NOT IN ('superadmin','admin','nachalnik')" + order, (shop,))
    elif role == 'master':
        cursor.execute(base_select + "WHERE u.shop_name = ? AND u.master_name = ? AND u.role NOT IN ('superadmin','admin','nachalnik','master')" + order, (shop, name))
    elif role == 'brigadir':
        cursor.execute(base_select + "WHERE u.shop_name = ? AND u.master_name = ? AND u.role = 'worker'" + order, (shop, name))
    else:
        conn.close()
        return []

    rows = cursor.fetchall()
    conn.close()
    return [dict(r) for r in rows]

def get_management_chain(user_id: int):
    """
    Returns the management hierarchy visible to a worker:
    nachalnik → master → brigadir of the same shop.
    Admin/superadmin are NEVER shown.
    """
    user = get_user(user_id)
    if not user:
        return []
    shop = user.get('shop_name', '')
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
    SELECT telegram_id, full_name, username, role, shop_name
    FROM users
    WHERE shop_name = ? AND role IN ('nachalnik','master','brigadir')
    ORDER BY
        CASE role
            WHEN 'nachalnik' THEN 1
            WHEN 'master'    THEN 2
            WHEN 'brigadir'  THEN 3
        END
    """, (shop,))
    rows = cursor.fetchall()
    conn.close()
    return [dict(r) for r in rows]

def get_attack_full_info(telegram_id: int):
    """Returns full profile of a user who attempted an attack, for admin notification."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
    SELECT u.telegram_id, u.full_name, u.username, u.phone,
           u.shop_name, u.master_name, u.role, u.invite_code,
           u.registered_at,
           COUNT(a.id) as total_attacks,
           MAX(a.attack_time) as last_attack,
           GROUP_CONCAT(a.attempt_details, ' | ') as all_attempts
    FROM users u
    LEFT JOIN attacks a ON a.telegram_id = u.telegram_id
    WHERE u.telegram_id = ?
    GROUP BY u.telegram_id
    """, (telegram_id,))
    row = cursor.fetchone()
    conn.close()
    return dict(row) if row else None

def get_all_admins():
    """Returns all admin/superadmin accounts."""
    import config
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
    SELECT telegram_id, full_name, username, role, permissions
    FROM users
    WHERE role IN ('admin', 'superadmin') OR telegram_id IN ({})
    """.format(','.join('?' for _ in config.ADMIN_IDS)), config.ADMIN_IDS)
    rows = cursor.fetchall()
    conn.close()
    admins = [dict(r) for r in rows]
    for a in admins:
        if a["telegram_id"] in config.ADMIN_IDS:
            a["role"] = "superadmin"
            a["permissions"] = "all"
    return admins

def get_all_management():
    """Returns all nachalnik/master/brigadir/quality/director/engineer for admin panel overview."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
    SELECT u.telegram_id, u.full_name, u.username, u.shop_name, u.master_name, u.role,
           COALESCE(MAX(tr.percentage), 0) as best_score,
           COUNT(tr.id) as tests_completed
    FROM users u
    LEFT JOIN test_results tr ON u.telegram_id = tr.user_telegram_id
    WHERE u.role IN ('nachalnik','master','brigadir','quality','director','engineer')
    GROUP BY u.telegram_id
    ORDER BY
        CASE u.role
            WHEN 'director'  THEN 1
            WHEN 'quality'   THEN 2
            WHEN 'engineer'  THEN 3
            WHEN 'nachalnik' THEN 4
            WHEN 'master'    THEN 5
            WHEN 'brigadir'  THEN 6
        END,
        u.shop_name
    """)
    rows = cursor.fetchall()
    conn.close()
    return [dict(r) for r in rows]

def set_user_role_and_permissions(telegram_id: int, role: str, permissions):
    conn = get_db()
    cursor = conn.cursor()
    perm_str = ",".join(permissions) if isinstance(permissions, (list, set)) else str(permissions or "")
    cursor.execute("UPDATE users SET role = ?, permissions = ? WHERE telegram_id = ?", (role, perm_str, telegram_id))
    conn.commit()
    conn.close()

def get_user_by_username_or_id(identifier: str):
    conn = get_db()
    cursor = conn.cursor()
    identifier = identifier.strip().lstrip("@")
    if identifier.isdigit():
        cursor.execute("SELECT * FROM users WHERE telegram_id = ?", (int(identifier),))
    else:
        cursor.execute("SELECT * FROM users WHERE LOWER(username) = LOWER(?)", (identifier.lower(),))
    row = cursor.fetchone()
    conn.close()
    return dict(row) if row else None

def is_admin_or_superadmin(telegram_id: int):
    import config
    if telegram_id in config.ADMIN_IDS:
        return True
    user = get_user(telegram_id)
    if user and user.get("role") in ("admin", "superadmin"):
        return True
    return False

def check_user_permission(telegram_id: int, permission: str):
    import config
    if telegram_id in config.ADMIN_IDS:
        return True
    user = get_user(telegram_id)
    if not user:
        return False
    if user.get("role") == "superadmin":
        return True
    if user.get("role") == "admin":
        perms = (user.get("permissions") or "").split(",")
        return permission in perms or "all" in perms or "*" in perms
    return False

# Test Result functions
def save_test_result(telegram_id: int, score: int, total: int, percentage: float, time_taken: int, mistakes: str = ""):
    cleanup_old_test_results()
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
    INSERT INTO test_results (user_telegram_id, score, total_questions, percentage, time_taken_seconds, mistakes)
    VALUES (?, ?, ?, ?, ?, ?)
    """, (telegram_id, score, total, percentage, time_taken, mistakes))
    conn.commit()
    conn.close()

def check_user_cooldown(telegram_id: int):
    """
    Checks if user failed their last test (<80%) and enforces a 5-minute (300s) retry cooldown.
    Returns (can_take: bool, remaining_seconds: int).
    """
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
    SELECT percentage, completed_at
    FROM test_results
    WHERE user_telegram_id = ?
    ORDER BY completed_at DESC LIMIT 1
    """, (telegram_id,))
    row = cursor.fetchone()
    conn.close()

    if not row:
        return True, 0

    last_pct = row["percentage"]
    if last_pct >= 80:
        return True, 0

    last_time_str = row["completed_at"]
    if not last_time_str:
        return True, 0

    try:
        if '.' in last_time_str:
            last_dt = datetime.datetime.strptime(last_time_str.split('.')[0], "%Y-%m-%d %H:%M:%S")
        else:
            last_dt = datetime.datetime.strptime(last_time_str, "%Y-%m-%d %H:%M:%S")
        
        now = datetime.datetime.utcnow()
        elapsed = (now - last_dt).total_seconds()
        cooldown_period = 300  # 5 minutes

        if elapsed < cooldown_period:
            remaining = int(cooldown_period - elapsed)
            return False, max(remaining, 1)
    except Exception as e:
        print(f"Error checking cooldown: {e}")

    return True, 0

def get_user_stats(telegram_id: int):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
    SELECT COUNT(*) as tests_count, MAX(percentage) as best_score, AVG(percentage) as avg_score
    FROM test_results WHERE user_telegram_id = ?
    """, (telegram_id,))
    row = cursor.fetchone()
    conn.close()
    return dict(row) if row else {"tests_count": 0, "best_score": 0, "avg_score": 0}

def get_leaderboard(limit: int = 500, shop_name: str = None):
    conn = get_db()
    cursor = conn.cursor()
    query = """
    SELECT u.telegram_id, u.full_name, u.username, u.phone, u.shop_name, u.master_name, u.role,
           COALESCE(MAX(tr.percentage), 0) as best_score, 
           COUNT(tr.id) as total_attempts,
           MAX(tr.completed_at) as last_test
    FROM users u
    LEFT JOIN test_results tr ON u.telegram_id = tr.user_telegram_id
    """
    params = []
    if shop_name and shop_name != "all":
        query += " WHERE u.shop_name = ?"
        params.append(shop_name)

    query += """
    GROUP BY u.telegram_id
    ORDER BY best_score DESC, total_attempts DESC, u.registered_at ASC
    LIMIT ?
    """
    params.append(limit)
    cursor.execute(query, tuple(params))
    rows = cursor.fetchall()
    conn.close()
    return [dict(r) for r in rows]

def get_all_workers_admin():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
    SELECT u.telegram_id, u.full_name, u.username, u.phone, u.shop_name, u.master_name, u.invite_code,
           COALESCE(MAX(tr.percentage), 0) as best_score,
           COUNT(tr.id) as tests_completed,
           (SELECT mistakes FROM test_results tr2 WHERE tr2.user_telegram_id = u.telegram_id ORDER BY completed_at DESC LIMIT 1) as latest_mistakes
    FROM users u
    LEFT JOIN test_results tr ON u.telegram_id = tr.user_telegram_id
    GROUP BY u.telegram_id
    ORDER BY best_score DESC
    """)
    rows = cursor.fetchall()
    conn.close()
    return [dict(r) for r in rows]

def get_shop_statistics():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
    SELECT 
        s.shop_name,
        COUNT(u.telegram_id) as total_workers,
        COUNT(user_best.best_score) as tested_workers,
        COALESCE(ROUND(AVG(user_best.best_score), 1), 0) as avg_score,
        COALESCE(SUM(CASE WHEN user_best.best_score >= 80 THEN 1 ELSE 0 END), 0) as expert_count
    FROM (
        SELECT shop_name FROM users WHERE shop_name IS NOT NULL AND shop_name != ''
        UNION
        SELECT shop_name FROM invite_codes WHERE shop_name IS NOT NULL AND shop_name != ''
    ) s
    LEFT JOIN users u ON u.shop_name = s.shop_name AND (u.role NOT IN ('superadmin', 'admin') OR u.role IS NULL)
    LEFT JOIN (
        SELECT user_telegram_id, MAX(percentage) as best_score
        FROM test_results
        GROUP BY user_telegram_id
    ) user_best ON u.telegram_id = user_best.user_telegram_id
    GROUP BY s.shop_name
    ORDER BY avg_score DESC, total_workers DESC
    """)
    rows = cursor.fetchall()
    conn.close()
    return [dict(r) for r in rows]

# ─────────────────────────────────────────────────────────────────
# Monthly Reports & Historical Archives Functions
# ─────────────────────────────────────────────────────────────────
MONTH_NAMES_RU = {
    "01": "Январь", "02": "Февраль", "03": "Март", "04": "Апрель",
    "05": "Май", "06": "Июнь", "07": "Июль", "08": "Август",
    "09": "Сентябрь", "10": "Октябрь", "11": "Ноябрь", "12": "Декабрь"
}
MONTH_NAMES_UZ = {
    "01": "Yanvar", "02": "Fevral", "03": "Mart", "04": "Aprel",
    "05": "May", "06": "Iyun", "07": "Iyul", "08": "Avgust",
    "09": "Sentabr", "10": "Oktabr", "11": "Noyabr", "12": "Dekabr"
}

def get_available_report_months():
    """
    Returns list of available months in DB format plus current month.
    Each item: {"code": "YYYY-MM", "label_ru": "Август 2026", "label_uz": "Avgust 2026", "count": int}
    """
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT strftime('%Y-%m', completed_at) as month_code, COUNT(*) as cnt
        FROM test_results
        WHERE completed_at IS NOT NULL
        GROUP BY month_code
        ORDER BY month_code DESC
    """)
    rows = cursor.fetchall()
    conn.close()

    months_dict = {}
    for r in rows:
        m_code = r["month_code"]
        if m_code:
            months_dict[m_code] = r["cnt"]

    # Always ensure current month is in list
    now_code = datetime.datetime.now().strftime("%Y-%m")
    if now_code not in months_dict:
        months_dict[now_code] = 0

    result = []
    for m_code in sorted(months_dict.keys(), reverse=True):
        parts = m_code.split("-")
        if len(parts) == 2:
            year, m_num = parts[0], parts[1]
            ru_name = MONTH_NAMES_RU.get(m_num, m_num)
            uz_name = MONTH_NAMES_UZ.get(m_num, m_num)
            result.append({
                "code": m_code,
                "label_ru": f"{ru_name} {year}",
                "label_uz": f"{uz_name} {year}",
                "count": months_dict[m_code]
            })
    return result

def get_monthly_test_results(month_code: str = None, shop_name: str = None):
    """
    Returns test statistics and results per worker for a specific month (YYYY-MM) or 'all'.
    """
    conn = get_db()
    cursor = conn.cursor()

    query = """
    SELECT u.telegram_id, u.full_name, u.username, u.phone, u.shop_name, u.master_name, u.role,
           COALESCE(MAX(tr.percentage), 0) as best_score,
           COALESCE(ROUND(AVG(tr.percentage), 1), 0) as avg_score,
           COUNT(tr.id) as tests_completed,
           MAX(tr.completed_at) as last_test_date,
           (SELECT mistakes FROM test_results tr2 
            WHERE tr2.user_telegram_id = u.telegram_id 
    """
    params = []
    if month_code and month_code != "all":
        query += " AND strftime('%Y-%m', tr2.completed_at) = ? "
        params.append(month_code)

    query += """ ORDER BY completed_at DESC LIMIT 1) as latest_mistakes
    FROM users u
    """

    if month_code and month_code != "all":
        query += " JOIN test_results tr ON u.telegram_id = tr.user_telegram_id AND strftime('%Y-%m', tr.completed_at) = ? "
        params.append(month_code)
    else:
        query += " LEFT JOIN test_results tr ON u.telegram_id = tr.user_telegram_id "

    where_clauses = []
    if shop_name and shop_name != "all":
        where_clauses.append("u.shop_name = ?")
        params.append(shop_name)

    if where_clauses:
        query += " WHERE " + " AND ".join(where_clauses)

    query += """
    GROUP BY u.telegram_id
    ORDER BY best_score DESC, tests_completed DESC, u.registered_at ASC
    """

    cursor.execute(query, tuple(params))
    rows = cursor.fetchall()
    conn.close()
    return [dict(r) for r in rows]

def generate_monthly_report_file(month_code: str = None, shop_name: str = None, file_format: str = "xlsx") -> str:
    """
    Generates Excel (.xlsx) or CSV file with the monthly test report.
    Returns absolute file path.
    """
    import os
    import json

    results = get_monthly_test_results(month_code=month_code, shop_name=shop_name)

    if not month_code or month_code == "all":
        month_label_ru = "За всё время"
        file_suffix = "ALL"
    else:
        parts = month_code.split("-")
        if len(parts) == 2:
            year, m_num = parts[0], parts[1]
            month_label_ru = f"{MONTH_NAMES_RU.get(m_num, m_num)} {year}"
            file_suffix = month_code
        else:
            month_label_ru = month_code
            file_suffix = month_code

    reports_dir = os.path.join(os.path.dirname(__file__), "reports")
    os.makedirs(reports_dir, exist_ok=True)

    filename = f"BIQS_Report_{file_suffix}.{file_format}"
    filepath = os.path.join(reports_dir, filename)

    role_map = {
        'superadmin': 'Суперадминистратор',
        'admin': 'Администратор',
        'director': 'Руководство',
        'quality': 'Контроль качества',
        'engineer': 'Инженер',
        'nachalnik': 'Начальник цеха',
        'master': 'Мастер участка',
        'brigadir': 'Бригадир',
        'worker': 'Рабочий'
    }

    if file_format == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Отчет BIQS"

            ws.views.sheetView[0].showGridLines = True

            ws.merge_cells("A1:K1")
            ws["A1"] = "СП УЗ ТОНГ ХОНГ КО — ЕЖЕМЕСЯЧНЫЙ ОТЧЕТ ПО ТЕСТИРОВАНИЮ BIQS"
            ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
            ws["A1"].fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 35

            ws.merge_cells("A2:K2")
            ws["A2"] = f"Период: {month_label_ru} | Цех: {shop_name or 'Все цеха'} | Дата экспорта: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}"
            ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="555555")
            ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[2].height = 20

            headers = [
                "№", "ФИО сотрудника", "Telegram ID / Username", "Телефон", 
                "Должность", "Цех / Участок", "Лучший результат (%)", 
                "Средний балл (%)", "Кол-во тестов", "Статус BIQS", 
                "Допущенные ошибки"
            ]

            ws.append([])
            ws.append(headers)
            ws.row_dimensions[4].height = 26

            header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            thin_border = Border(
                left=Side(style='thin', color='D9D9D9'),
                right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'),
                bottom=Side(style='thin', color='D9D9D9')
            )

            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=4, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            pass_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            fail_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            pass_font = Font(name="Calibri", size=10, bold=True, color="276A3C")
            fail_font = Font(name="Calibri", size=10, bold=True, color="C65911")

            for idx, r in enumerate(results, 1):
                best = round(r.get("best_score", 0))
                avg = r.get("avg_score", 0)
                passed = best >= 80
                status = "СДАЛ (Эксперт BIQS)" if passed else "НЕ СДАЛ (<80%)"
                
                user_tg = f"@{r['username']}" if r.get('username') else str(r.get('telegram_id', ''))
                role_title = role_map.get(r.get('role'), r.get('role', '—'))

                mistakes_str = "—"
                if r.get('latest_mistakes'):
                    try:
                        ml = json.loads(r['latest_mistakes'])
                        mistakes_str = ", ".join(ml) if isinstance(ml, list) else str(r['latest_mistakes'])
                    except:
                        mistakes_str = str(r['latest_mistakes'])

                row_values = [
                    idx,
                    r.get("full_name", "—"),
                    user_tg,
                    r.get("phone") or "—",
                    role_title,
                    r.get("shop_name") or "—",
                    f"{best}%",
                    f"{avg}%",
                    r.get("tests_completed", 0),
                    status,
                    mistakes_str
                ]

                row_idx = idx + 4
                ws.append(row_values)
                ws.row_dimensions[row_idx].height = 22

                for col_num in range(1, len(row_values) + 1):
                    cell = ws.cell(row=row_idx, column=col_num)
                    cell.border = thin_border
                    cell.font = Font(name="Calibri", size=10)
                    if col_num in (1, 7, 8, 9):
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif col_num == 10:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.fill = pass_fill if passed else fail_fill
                        cell.font = pass_font if passed else fail_font
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")

            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
            ws.column_dimensions['B'].width = 28
            ws.column_dimensions['K'].width = 35

            wb.save(filepath)
            return filepath
        except Exception as e:
            print(f"[EXCEL ERROR] Failed openpyxl export: {e}. Falling back to CSV.")

    import csv
    csv_path = filepath.replace(".xlsx", ".csv")
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(["СП УЗ ТОНГ ХОНГ КО — ЕЖЕМЕСЯЧНЫЙ ОТЧЕТ ПО ТЕСТИРОВАНИЮ BIQS"])
        writer.writerow([f"Период: {month_label_ru} | Цех: {shop_name or 'Все цеха'}"])
        writer.writerow([])
        writer.writerow(["№", "ФИО сотрудника", "Telegram ID / Username", "Телефон", "Должность", "Цех / Участок", "Лучший результат (%)", "Средний балл (%)", "Кол-во тестов", "Статус BIQS", "Допущенные ошибки"])
        
        for idx, r in enumerate(results, 1):
            best = round(r.get("best_score", 0))
            avg = r.get("avg_score", 0)
            status = "СДАЛ (80%+)" if best >= 80 else "НЕ СДАЛ (<80%)"
            user_tg = f"@{r['username']}" if r.get('username') else str(r.get('telegram_id', ''))
            role_title = role_map.get(r.get('role'), r.get('role', '—'))
            mistakes_str = "—"
            if r.get('latest_mistakes'):
                try:
                    ml = json.loads(r['latest_mistakes'])
                    mistakes_str = ", ".join(ml) if isinstance(ml, list) else str(r['latest_mistakes'])
                except:
                    mistakes_str = str(r['latest_mistakes'])

            writer.writerow([
                idx, r.get("full_name","—"), user_tg, r.get("phone") or "—",
                role_title, r.get("shop_name") or "—", f"{best}%", f"{avg}%",
                r.get("tests_completed", 0), status, mistakes_str
            ])
    return csv_path





def add_banned_prompt(pattern: str, category: str = 'general'):
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO banned_prompts (pattern, category) VALUES (?, ?)", (pattern.strip().lower(), category))
        conn.commit()
        success = True
    except Exception:
        success = False
    conn.close()
    return success

def get_all_banned_prompts():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT id, pattern, category, added_at FROM banned_prompts ORDER BY category, pattern")
    rows = cursor.fetchall()
    conn.close()
    return [dict(r) for r in rows]

def delete_banned_prompt(pattern_id: int):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM banned_prompts WHERE id = ?", (pattern_id,))
    deleted = cursor.rowcount > 0
    conn.commit()
    conn.close()
    return deleted

def check_security_violations(text: str, user_id: int = None):
    """
    Checks text against security patterns (profanity, links, prompt injection attacks).
    Returns (is_violation: bool, violation_title: str, matched_pattern: str)
    """
    if not text:
        return False, None, None

    # Superadmin / Admin bypass
    if user_id and is_admin_or_superadmin(user_id):
        return False, None, None

    text_lower = text.lower().strip()

    # Regex check for Telegram Bot Token leak / injection pattern
    import re
    if re.search(r'[0-9]{8,10}:[A-Za-z0-9_-]{35}', text):
        return True, 'Утечка / Попытка внедрения Telegram Bot Token', 'Bot Token Regex ([0-9]{8..}:[A-Za-z0-9_-]{35})'

    banned_items = get_all_banned_prompts()
    for item in banned_items:
        pat = item['pattern'].lower()
        cat = item['category']

        if pat in text_lower:
            type_labels = {
                'profanity': 'Ненормативная лексика / Матерные слова',
                'link': 'Несанкционированная ссылка',
                'prompt_injection': 'Вирусный промпт / Попытка взлома',
                'sql_injection': 'SQL-Инъекция',
                'xss': 'XSS-Атака',
                'code_injection': 'Инъекция кода / Вредоносный скрипт',
                'command_injection': 'Системная команда',
                'honeypot': 'Ловушка / Попытка поиска Админ-панели (Honeypot)',
                'attack_pattern': 'Вредоносный паттерн атаки'
            }
            violation_title = type_labels.get(cat, 'Запрещенный промпт')
            return True, violation_title, pat

    return False, None, None


def log_attack(telegram_id: int, attempt_details: str):

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
    INSERT INTO attacks (telegram_id, attempt_details)
    VALUES (?, ?)
    """, (telegram_id, attempt_details))
    conn.commit()
    conn.close()

def get_attacks_summary():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
    SELECT u.full_name, u.shop_name, u.master_name, COUNT(a.id) as attack_count
    FROM attacks a
    JOIN users u ON a.telegram_id = u.telegram_id
    GROUP BY a.telegram_id
    ORDER BY attack_count DESC
    """)
    rows = cursor.fetchall()
    conn.close()
    return [dict(r) for r in rows]

def get_attacks_detailed():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
    SELECT a.telegram_id, a.attempt_details, a.attack_time, u.full_name, u.shop_name, u.master_name, u.username, u.phone
    FROM attacks a
    JOIN users u ON a.telegram_id = u.telegram_id
    ORDER BY a.attack_time DESC
    LIMIT 50
    """)
    rows = cursor.fetchall()
    conn.close()
    return [dict(r) for r in rows]

# Data Provider for Official 30 BIQS Elements
def get_biqs_elements():
    return [
        {
            "id": 1,
            "code": "BIQS-01",
            "title_uz": "Nomuvofiq mahsulot bilan ishlash yo'riqnomasi",
            "title_ru": "Изоляция и контроль несоответствующей продукции",
            "desc_uz": "Сифат таъминлаш стандарти ва номувофиқ маҳсулот учун ҳар бир иш жойида махсус қизил рангли идишли жой, ёрлиқ ва назорат ҳужжати мавжудлиги.",
            "desc_ru": "Стандарт качества и работа с браком. Наличие красной зоны (Red Tag), ярлыков и документации контроля на каждом рабочем месте.",
            "icon": "🚫"
        },
        {
            "id": 2,
            "code": "BIQS-02",
            "title_uz": "Ko'p pog'onali audit (Layered Audit - LA)",
            "title_ru": "Многоуровневый аудит процессов (LA)",
            "desc_uz": "Кўп поғонали аудит йўриқномаси. Ҳар бир жамоа хатар саволномаси асосида LA вароғини юритиши ва ҳафталик ҳисоботни раҳбариятга топшириши.",
            "desc_ru": "Инструкция многоуровневого аудита. Ведение листов LA на участках и предоставление еженедельных отчетов руководству.",
            "icon": "📝"
        },
        {
            "id": 3,
            "code": "BIQS-03",
            "title_uz": "PFMEA режаси ва хатарларни баҳолаш",
            "title_ru": "Анализ рисков и планирование PFMEA",
            "desc_uz": "PFMEA командасини тузиш, жараён муаммоларини тўлиқ киритиш. Мезон ва баллга асосан чора-тадбирлар кўриш ва хавф баллини тушириш.",
            "desc_ru": "Формирование команды PFMEA, полный учет рисков процесса, разработка мероприятий и снижение балла риска.",
            "icon": "⚖️"
        },
        {
            "id": 4,
            "code": "BIQS-04",
            "title_uz": "E'tirozlar ва PFMEA хатарларини кўриб чиқиш",
            "title_ru": "Анализ претензий (GCA/DRR) в PFMEA",
            "desc_uz": "Иш жараёнидаги камчиликлар ва истеъмолчи томонидан келган эътирозларни (GCA, DRR, Reclamation) PFMEA рискларида тўлиқ кўриб чиқиш.",
            "desc_ru": "Учет всех выявленных дефектов и рекламаций клиентов (GCA, DRR) в структуре рисков PFMEA.",
            "icon": "📢"
        },
        {
            "id": 5,
            "code": "BIQS-05",
            "title_uz": "Aylanib o'tish boshqaruvi (Bypass Management)",
            "title_ru": "Управление обходными технологиями (Bypass)",
            "desc_uz": "Bypass Management методик қўлланмалари асосида стандарт йўриқномалар яратиш, ходимларни ўқитиш ва ВРМ жараёнларига киритиш.",
            "desc_ru": "Создание стандартных инструкций по Bypass Management, обучение персонала и внедрение во все процессы ВРМ.",
            "icon": "🔀"
        },
        {
            "id": 6,
            "code": "BIQS-06",
            "title_uz": "Xatolardan xoli qilish pog'onalari (Poka-Yoke)",
            "title_ru": "Защита от ошибок (Poka-Yoke / Check Fix)",
            "desc_uz": "1-поғона: Маркерлар. 2-поғона: Инструментлар (Check Fix/Torque). 3-поғона: Функционал NG детал текшируви. 4-поғона: ТРМ текшируви.",
            "desc_ru": "1-уровень: Маркировка деталей. 2-уровень: Инструменты (Check Fix/Torque). 3-уровень: Проверка NG деталей. 4-уровень: Контроль TPM.",
            "icon": "🛡️"
        },
        {
            "id": 7,
            "code": "BIQS-07",
            "title_uz": "O'lchov vositalari va MSA tahlili",
            "title_ru": "Поверка приборов и анализ MSA",
            "desc_uz": "Барча текширув инструментлари ва жиҳозлар поверкадан ўтганлиги, ёрлиғи борлиги ва MSA махсус таҳлилини ўтказиш.",
            "desc_ru": "Своевременная поверка всех измерительных приборов, наличие действующей маркировки и проведение анализа MSA.",
            "icon": "📏"
        },
        {
            "id": 8,
            "code": "BIQS-08",
            "title_uz": "Tezkor munosabat жараёни (Fast Response)",
            "title_ru": "Процесс оперативного реагирования (Fast Response)",
            "desc_uz": "Лимитдан ошган муаммоларни Fast Response жараёнига олиб чиқиш, ҳужжатлаштириш ва бартараф этилишини назорат қилиш.",
            "desc_ru": "Вынос превышающих лимит проблем на процесс Fast Response, документирование и контроль их устранения.",
            "icon": "⚡"
        },
        {
            "id": 9,
            "code": "BIQS-09",
            "title_uz": "Муаммоларни ҳал этиш (PPSR & Escalation)",
            "title_ru": "Решение проблем PPSR и эскалация",
            "desc_uz": "PPSR жараёнига барча алоқадорларни жалб этиш, эскалация орқали қўшимча чора-тадбирлар ва ўқитишни тақдим этиш.",
            "desc_ru": "Вовлечение команды в процесс PPSR, эскалация проблем, проведение дополнительных проверок и обучения.",
            "icon": "🔍"
        },
        {
            "id": 10,
            "code": "BIQS-10",
            "title_uz": "Сифат текшируви ҳужжатлари",
            "title_ru": "Документирование контроля качества",
            "desc_uz": "Сифат текширув жараёнини белгиланган талаб даражасида ҳужжатлаштириш, маълумот ҳолатини назорат қилиш ва чоралар кўриш.",
            "desc_ru": "Документирование процессов контроля качества, мониторинг данных и оперативное принятие коррективных мер.",
            "icon": "📄"
        },
        {
            "id": 11,
            "code": "BIQS-11",
            "title_uz": "Standartlashtirilgan ish (SOS / JES)",
            "title_ru": "Стандартизированная работа (SOS / JES)",
            "desc_uz": "Цикли ва ноцикли ишлар хавфсизлик, сифат ва вақт талаблари асосида SOS форматида ҳужжатлаштирилган ва станцияда кўрсатилган.",
            "desc_ru": "Документирование операций в формате SOS с указанием требований безопасности, качества и времени.",
            "icon": "📋"
        },
        {
            "id": 12,
            "code": "BIQS-12",
            "title_uz": "4M o'zgarishlar назорати (4M Control)",
            "title_ru": "Управление изменениями 4M (Man, Machine, Material, Method)",
            "desc_uz": "Одам, Жиҳоз, Машина ва Жараён (4М) бўйича ўзгаришларни Breakpoint ва PFMEA орқали назорат қилиш ва тасдиқлаш.",
            "desc_ru": "Контроль изменений по 4M (Человек, Оборудование, Материал, Процесс) через точки Breakpoint и PFMEA.",
            "icon": "🔄"
        },
        {
            "id": 13,
            "code": "BIQS-13",
            "title_uz": "Сифат текширув воситалари ва самарадорлиги",
            "title_ru": "Эффективность средств контроля качества",
            "desc_uz": "Сифат текширувлари стандарт ҳужжатларда кўрсатилганлиги, ходимлар ўлчашни билиши ва самарадорлиги кўриб чиқилиши.",
            "desc_ru": "Наличие всех средств контроля по стандартам, обучение сотрудников методам измерений и оценка эффективности проверок.",
            "icon": "🔍"
        },
        {
            "id": 14,
            "code": "BIQS-14",
            "title_uz": "O'zgarishlar бланкаси ва Breakpoint (PTR)",
            "title_ru": "Оформление PTR и контроль Breakpoint",
            "desc_uz": "Ҳар бир детал ва жараён ўзгариши учун PTR бланкаси тўлдирилиши, Breakpoint назорат қилиниши ва 4М доскасида акс эттирилиши.",
            "desc_ru": "Заполнение бланков PTR при любых изменениях деталей и процессов, фиксирование точек Breakpoint на досках 4M.",
            "icon": "🏷️"
        },
        {
            "id": 15,
            "code": "BIQS-15",
            "title_uz": "Andon ogohlantirish tizimi (Andon System)",
            "title_ru": "Система оповещения Andon",
            "desc_uz": "АНДОН тизими барча участкаларда ишлайди. Муаммо аниқланганда хабар бериш, маълумотларни таҳлил қилиш ва реакция ҳужжатлаштирилган.",
            "desc_ru": "Функционирование системы Andon на всех участках, сбор данных и зафиксированная реакция на вызовы.",
            "icon": "🚨"
        },
        {
            "id": 16,
            "code": "BIQS-16",
            "title_uz": "Муаммолар эскалацияси ва лимитлари",
            "title_ru": "Процесс эскалации проблем",
            "desc_uz": "Муаммо мезонлари ва эскалация жараёнини барча ходимлар билиши. Лимитдан ошганда тегишли чора-тадбирлар кўрсатилиши.",
            "desc_ru": "Знание сотрудниками критериев проблем и регламента эскалации. Принятие мер при превышении установленных лимитов.",
            "icon": "📈"
        },
        {
            "id": 17,
            "code": "BIQS-17",
            "title_uz": "Визуал бошқарув ва иш жойи стандарти",
            "title_ru": "Стандарты визуального менеджмента",
            "desc_uz": "Иш жойини ташкил этиш ва визуал бошқарув стандартлари аниқ. Махсус қўлланмалар NG ва OK ҳолатида кўрсатилган.",
            "desc_ru": "Внедрение визуального менеджмента по всему предприятию. Наглядные инструкции с разделением состояний OK и NG.",
            "icon": "👁️"
        },
        {
            "id": 18,
            "code": "BIQS-18",
            "title_uz": "Визуал йўриқномалар ва ўқитиш",
            "title_ru": "Визуальные инструкции и обучение",
            "desc_uz": "Сифатга оид йўриқномалар ходимларга ўргатилган ва қулай ўрнатилган. Ўзгаришлар самарали етказилади.",
            "desc_ru": "Доведение визуальных инструкций по качеству до сотрудников, удобное размещение на рабочих местах и актуализация.",
            "icon": "🖼️"
        },
        {
            "id": 19,
            "code": "BIQS-19",
            "title_uz": "Жараён назорати ҳужжатлари (Control Plan)",
            "title_ru": "План контроля процессов (Flowchart, FMEA, CP)",
            "desc_uz": "Ҳар бир технологик жараён учун CP, FLOWCHART, FMEA, SOS/JES бўлиши ва улар бир-бирига тўлиқ мос келиши.",
            "desc_ru": "Наличие документации CP, Flowchart, FMEA, SOS/JES для каждого техпроцесса и их полное взаимное соответствие.",
            "icon": "📊"
        },
        {
            "id": 20,
            "code": "BIQS-20",
            "title_uz": "Ҳужжатлар мослиги ва CheckList",
            "title_ru": "Проверка соответствия документов и чек-листов",
            "desc_uz": "CP, Flowchart, FMEA, SOS, CheckList амалда бир-бирига мослигини ва иш жараёни ҳужжат асосида кечишини текшириш.",
            "desc_ru": "Проверка соответствия реального техпроцесса документам CP, FMEA, SOS и чек-листам на рабочих местах.",
            "icon": "✅"
        },
        {
            "id": 21,
            "code": "BIQS-21",
            "title_uz": "Критик ва хавфли нуқталар (QCOS Torque / Weld)",
            "title_ru": "Контроль критических точек (QCOS Torque/Weld)",
            "desc_uz": "Критик ва хавфли технологик нуқталар (QCOS Torque/Weld Check) белгиланган, муддатли текширув ва SPC таҳлили ўтказилади.",
            "desc_ru": "Определение и регулярный контроль критических точек (QCOS Torque/Weld), проведение статистического анализа SPC.",
            "icon": "⚙️"
        },
        {
            "id": 22,
            "code": "BIQS-22",
            "title_uz": "Таъмирлаш участкаси ва малака (Rework & Flexibility)",
            "title_ru": "Зона ремонта деталей и квалификация персонала",
            "desc_uz": "Таъмирлаш учун алоҳида жой, SOS/JES йўриқномалари ва ходимнинг Flexibility chart бўйича малакаси мавжудлиги.",
            "desc_ru": "Наличие изолированной зоны ремонта деталей, специальных инструкций SOS/JES и матрицы квалификации (Flexibility chart).",
            "icon": "🔧"
        },
        {
            "id": 23,
            "code": "BIQS-23",
            "title_uz": "Сифат муаммоларини хабар қилиш (Quality Alert)",
            "title_ru": "Оповещение о проблемах качества (Quality Alert)",
            "desc_uz": "Сифат муаммолари ҳақида олдинга ва орқага зудлик билан хабар бериш (Containment Sheet, Quality Alert) ва назорат қилиш.",
            "desc_ru": "Двустороннее оперативное оповещение о проблемах качества (Containment Sheet, Quality Alert) и контроль их устранения.",
            "icon": "🛎️"
        },
        {
            "id": 24,
            "code": "BIQS-24",
            "title_uz": "Ходимлар малакаси ва JIT ўқитиш",
            "title_ru": "Квалификация операторов и обучение JIT",
            "desc_uz": "Ҳар бир операция учун SOS/JES асосида JIT ўқитиш ва Flexibility chart орқали малака ва ўргатиш режасини назорат қилиш.",
            "desc_ru": "Обучение персонала по системе JIT на основе инструкций SOS/JES и контроль матрицы квалификации (Flexibility chart).",
            "icon": "🎓"
        },
        {
            "id": 25,
            "code": "BIQS-25",
            "title_uz": "Иш жойи ва маҳсулот софлиги (Cleanliness)",
            "title_ru": "Чистота рабочего места и защита от загрязнений",
            "desc_uz": "Иш жойи ва маҳсулотни чанг, кир ва ифлосланишдан ҳимоя қилиш, тозалаш йўриқномаси ва назорати.",
            "desc_ru": "Защита рабочего места и продукции от пыли и загрязнений, наличие инструкций по очистке и постоянный контроль.",
            "icon": "🧹"
        },
        {
            "id": 26,
            "code": "BIQS-26",
            "title_uz": "ТРМ ва ППР профилактика назорати",
            "title_ru": "Контроль обслуживания TPM и ремонта (ППР)",
            "desc_uz": "ТРМ чек-листлари, ППР режали таъмирлаш ҳужжатлари, эҳтиёт қисмлар мавжудлиги ва BIQS-1, 15 талаблари бажарилиши.",
            "desc_ru": "Ведение чек-листов TPM, проведение ППР, контроль наличия запчастей и выполнение требований BIQS-1, 15.",
            "icon": "🛠️"
        },
        {
            "id": 27,
            "code": "BIQS-27",
            "title_uz": "FIFO йўриқномаси ва ротация",
            "title_ru": "Соблюдение регламента FIFO и маркировка",
            "desc_uz": "FIFO йўриқномаси, полдаги FIFO йўналишлари, идишлардаги ёрлиқлар ва санага асосан кетма-кет маҳсулот олиниши.",
            "desc_ru": "Соблюдение принципа FIFO, наличие напольной разметки, ярлыков с датами и последовательное расходование материалов.",
            "icon": "⏳"
        },
        {
            "id": 28,
            "code": "BIQS-28",
            "title_uz": "Контейнер ва тара таъминоти (Approved Container)",
            "title_ru": "Контейнеризация, тара и соблюдение Min/Max",
            "desc_uz": "Маҳсулот учун тасдиқланган контейнер ва таралардан фойдаланиш, биркалар мавжудлиги ва Min/Max жараёнига амал қилиш.",
            "desc_ru": "Использование только утвержденной тары и контейнеров, наличие бирк и соблюдение лимитов запасов Min/Max.",
            "icon": "📦"
        },
        {
            "id": 29,
            "code": "BIQS-29",
            "title_uz": "Таъминотчилар сифати бошқаруви (SUB Quality)",
            "title_ru": "Управление качеством поставщиков (SUB)",
            "desc_uz": "Маҳаллийлаштириш ва SUB таъминоти кириш назорати (IQC), муаммоларни тезкор етказиш ва BIQS 1-13 бўйича текширувлар.",
            "desc_ru": "Входной контроль (IQC) компонентов локализации, оперативная претензионная работа и аудит поставщиков по BIQS 1-13.",
            "icon": "🚚"
        },
        {
            "id": 30,
            "code": "BIQS-30",
            "title_uz": "Меҳнат муҳофазаси ва хавфсизлик (SCOS & GMS PI)",
            "title_ru": "Инструкции по безопасности (SCOS) и GMS PI",
            "desc_uz": "Хавфсизлик бўйича йўриқномалар ва кўргазмали қўлланмалар (SCOS) бўлиши ҳамда GMS PI 2,3 талабларига пўлатдек амал қилиниши.",
            "desc_ru": "Наличие наглядных инструкций по безопасности (SCOS) и строгое выполнение требований GMS PI 2,3.",
            "icon": "🤯"
        }
    ]

def get_biqs_questions():
    return [
    {
        "id": 1,
        "question_uz": "Nomuvofiq (brak) mahsulot aniqlanganda har bir ish joyida nima bo'lishi va nima qilinishi shart (BIQS-01)?",
        "question_ru": "Что должно быть на каждом рабочем месте при обнаружении дефектной продукции (BIQS-01)?",
        "options_uz": [
            "Mahsulotni smena oxirida saralab, yashil zonaga joylashtirish",
            "Maxsus qizil rangli idish (Red Tag), yorliq va mahsulotni ajratish",
            "Operatordan so'ramay turib Rework zonasiga o'tkazish",
            "Mahsulotni nazorat daftarlariga yozib, liniyada qoldirish"
        ],
        "options_ru": [
            "Сортировка деталей в конце смены и размещение в Зеленой зоне",
            "Красная зона (Red Tag), ярлык и немедленная изоляция брака",
            "Передача детали на участок Rework без оформления бирки",
            "Запись дефекта в журнал контроля с оставлением детали на линии"
        ],
        "correct": 1,
        "explanation_uz": "BIQS-01: Har bir ish joyida qizil idish va yorliq bo'lishi hamda nuqson darhol ajratilishi shart!",
        "explanation_ru": "BIQS-01: Брак немедленно изолируется в красную зону (Red Tag) с оформлением ярлыка."
    },
    {
        "id": 2,
        "question_uz": "Ko'p pog'onali audit (LA - Layered Audit) bo'yicha haftalik hisobot kimga taqdim etiladi (BIQS-02)?",
        "question_ru": "Кому предоставляется еженедельный отчет по многоуровневому аудиту LA (BIQS-02)?",
        "options_uz": [
            "Faqat smena ustasiga va uchastka brigadiriga",
            "Yuqori rahbariyatga va sifat audit jamoasiga",
            "Tashqi sertifikatlashtirish organiga har oyda",
            "Ta'minotchilar bo'limiga va kirish nazoratiga"
        ],
        "options_ru": [
            "Только мастеру смены и бригадиру участка",
            "Высшему руководству завода и команде аудита",
            "Ежемесячно в внешний орган сертификации",
            "В отдел поставщиков и входного контроля"
        ],
        "correct": 1,
        "explanation_uz": "BIQS-02: LA monitoringi haftalik hisobot shaklida yuqori rahbariyatga taqdim etiladi.",
        "explanation_ru": "BIQS-02: Ведется мониторинг листов LA с еженедельным докладом высшему руководству."
    },
    {
        "id": 3,
        "question_uz": "PFMEA qachon qayta ko'rib chiqiladi va uning maqsadi nima (BIQS-03)?",
        "question_ru": "С какой целью проводится пересмотр PFMEA (BIQS-03)?",
        "options_uz": [
            "Faqat yangi uskunalar sotib olish uchun smeta tuzish",
            "Xatarlar ballini (Risk Score) tushirish va profilaktik choralarni belgilash",
            "Braklar sonini oylik maosh jadvaliga bog'lash",
            "Faqat iste'molchiga tayyor mahsulot yuborilgandan so'ng tahlil qilish"
        ],
        "options_ru": [
            "Составление сметы только для закупки нового оборудования",
            "Снижение балла риска (Risk Score) и назначение превентивных мер",
            "Привязка количества брака к ведомости заработной платы",
            "Анализ рисков только после отправки готовой продукции клиенту"
        ],
        "correct": 1,
        "explanation_uz": "BIQS-03: PFMEA orqali jarayon xatarlari baholanib, risk balli pasaytirilishi kerak.",
        "explanation_ru": "BIQS-03: PFMEA используется для анализа рисков и снижения балла вероятности дефекта."
    },
    {
        "id": 4,
        "question_uz": "Iste'molchidan kelgan e'tirozlar (GCA, DRR, Reklamatsiya) qaysi hujjatda ko'rib chiqilishi shart (BIQS-04)?",
        "question_ru": "Где обязательно должны рассматриваться претензии клиентов (GCA, DRR) (BIQS-04)?",
        "options_uz": [
            "Faqat ichki audit jurnali va 5S ko'rgazmalarida",
            "PFMEA xatarlar tahlili va jarayon xarita (Flowchart) hujjatlarida",
            "Faqat Fast Response taqvimiy yig'ilishida",
            "Standartlashtirish bo'limining yillik hisobotida"
        ],
        "options_ru": [
            "Только в журнале внутренних аудитов и стендах 5S",
            "В структуре рисков PFMEA и картах процесса Flowchart",
            "Исключительно на календарном собрании Fast Response",
            "В годовом отчете отдела стандартизации"
        ],
        "correct": 1,
        "explanation_uz": "BIQS-04: Iste'molchi e'tirozlari albatta PFMEA xatarlariga kiritilib tahlil qilinishi shart.",
        "explanation_ru": "BIQS-04: Все претензии потребителей должны быть включены в структуру рисков PFMEA."
    },
    {
        "id": 5,
        "question_uz": "Bypass Management (Aylanib o'tish) jarayoni qachon qo'llaniladi (BIQS-05)?",
        "question_ru": "При каких условиях применяется процесс Bypass Management (BIQS-05)?",
        "options_uz": [
            "Tezkorlikni oshirish uchun ustaning og'zaki ko'rsatmasi bilan",
            "Standart parametrlarni 20% ga o'zgartirgan holda avtomatik rejimda",
            "Standart yo'riqnoma (Bypass) va 100% vaqtinchalik qo'shimcha nazorat bilan",
            "Smena brigadirining ruxsati bilan 50% tanlanma nazorat ostida"
        ],
        "options_ru": [
            "Устным распоряжением мастера для ускорения сборки",
            "В автоматическом режиме с изменением допусков на 20%",
            "По утвержденной инструкции Bypass с введением 100% временного контроля",
            "С разрешения бригадира с выборочным контролем 50% деталей"
        ],
        "correct": 2,
        "explanation_uz": "BIQS-05: Bypass faqat standart yo'riqnoma va qo'shimcha nazorat ostida bajariladi.",
        "explanation_ru": "BIQS-05: Процесс Bypass требует специальных инструкций и 100% контроля."
    },
    {
        "id": 6,
        "question_uz": "Xatolardan xoli qilishning (Poka-Yoke) 2-pog'onasi nimani nazarda tutadi (BIQS-06)?",
        "question_ru": "Что подразумевает 2-й уровень защиты от ошибок в BIQS-06?",
        "options_uz": [
            "Vizual belgilash va detalga rangli marker qo'yish",
            "NG detali tushganda konveyerni avtomatik to'xtatuvchi sensor",
            "Tekshiruv instrumentlari (Check Fix/Torque) orqali parametrni qayd etish",
            "TPM profilaktik checklistini to'ldirish"
        ],
        "options_ru": [
            "Визуальная маркировка и нанесение цветного маркера",
            "Автоматическая блокировка конвейера датчиком при NG детали",
            "Инструментальный контроль параметров (Check Fix / Torque)",
            "Заполнение чек-листа профилактики TPM"
        ],
        "correct": 2,
        "explanation_uz": "BIQS-06: 2-pog'ona bu maxsus o'lchov instrumentlari (Torque, Check Fix) orqali sifatni kafolatlash.",
        "explanation_ru": "BIQS-06: 2-й уровень — это замер параметров приборами (Torque/Check Fix)."
    },
    {
        "id": 7,
        "question_uz": "O'lchov asboblari va jihozlarning to'g'riligini tasdiqlovchi hujjat (BIQS-07)?",
        "question_ru": "Что подтверждает точность измерительных приборов (BIQS-07)?",
        "options_uz": [
            "Zavod texnik pasporti va kalibrovka sertifikati",
            "Yaroqlilik yorlig'i va MSA (Gage R&R) tahlili bilan",
            "Smena ustasining muhrlangan tasdiqnomasi",
            "Visual Management doskasidagi asboblar ro'yxati"
        ],
        "options_ru": [
            "Заводской технический паспорт и сертификат калибровки",
            "Действующий ярлык поверки и проведенный анализ MSA (Gage R&R)",
            "Печать и подпись мастера смены в журнале",
            "Наличие прибора в реестре визуального менеджмента"
        ],
        "correct": 1,
        "explanation_uz": "BIQS-07: Barcha o'lchov vositalari poverkadan o'tganligi (yorliq) va MSA tahlili bo'lishi shart.",
        "explanation_ru": "BIQS-07: Все приборы должны иметь бирку о поверке и проходить анализ MSA."
    },
    {
        "id": 8,
        "question_uz": "Limitdan oshgan sifat muammolari qaysi jarayonga olib chiqiladi (BIQS-08)?",
        "question_ru": "На какой процесс выносятся проблемы качества, превысившие лимит (BIQS-08)?",
        "options_uz": [
            "Fast Response (Tezkor munosabat) kunlik yig'ilishiga",
            "Oylik sifat komissiyasi tahliliga",
            "Poka-Yoke 4-pog'ona auditi daqiqalariga",
            "SUB (Ta'minotchilar) sifat nazorati bo'limiga"
        ],
        "options_ru": [
            "На ежедневное совещание Fast Response (Оперативное реагирование)",
            "На ежемесячный анализ комиссии по качеству",
            "В протокол аудита Poka-Yoke 4-го уровня",
            "В отдел качества поставщиков SUB"
        ],
        "correct": 0,
        "explanation_uz": "BIQS-08: Katta muammolar zudlik bilan Fast Response yig'ilishida ko'rib chiqiladi.",
        "explanation_ru": "BIQS-08: Крупные дефекты незамедлительно выносятся на стенд Fast Response."
    },
    {
        "id": 9,
        "question_uz": "PPSR jarayonining asosiy maqsadi nima (BIQS-09)?",
        "question_ru": "Какова главная цель процесса PPSR (BIQS-09)?",
        "options_uz": [
            "Nuqsonli detal uchun mas'ul xodimga jarima belgilash",
            "Brak mahsulotni tezkor Rework zonasida qayta ishlash",
            "Jamoaviy ildiz sababni (Root Cause) topish va eskalatsiya choralari",
            "Uchastka SOS yo'riqnomasini bekor qilish"
        ],
        "options_ru": [
            "Наложение штрафа на оператора, допустившего брак",
            "Быстрая переработка бракованной продукции в зоне Rework",
            "Командный поиск корневой причины (Root Cause) и эскалация мер",
            "Отмена действующей инструкции SOS на участке"
        ],
        "correct": 2,
        "explanation_uz": "BIQS-09: PPSR orqali muammolar jamoaviy hal qilinadi va eskalatsiya qilinadi.",
        "explanation_ru": "BIQS-09: Процесс PPSR направлен на командное устранение причин дефекта."
    },
    {
        "id": 10,
        "question_uz": "Sifat tekshiruvi hujjatlashtirilishining asosiy talabi nima (BIQS-10)?",
        "question_ru": "Каково главное требование к документированию проверок качества (BIQS-10)?",
        "options_uz": [
            "Faqat smena yakunida umumiy sonni elektron bazaga kirish",
            "Standart blanqlarda doimiy qayd etish va chetlanishda chora ko'rish",
            "Faqat GCA va DRR tekshiruvlarida qayd etish",
            "Visual Management taxtasiga tahminiy grafik chizish"
        ],
        "options_ru": [
            "Внесение общего количества деталей в базу только в конце смены",
            "Регулярная запись в стандартизированных бланках с реакцией на отклонения",
            "Фиксация данных исключительно при проверках GCA и DRR",
            "Нанесение примерного графика на доску визуального менеджмента"
        ],
        "correct": 1,
        "explanation_uz": "BIQS-10: Barcha sifat tekshiruvlari rasmiy tasdiqlangan hujjatlarga qayd qilinishi kerak.",
        "explanation_ru": "BIQS-10: Результаты проверок фиксируются в официальных бланках."
    },
    {
        "id": 11,
        "question_uz": "Standartlashtirilgan ish (SOS/JES) o'z ichiga nimalarni qamrab oladi (BIQS-11)?",
        "question_ru": "Что охватывает стандартизированная работа (SOS/JES) (BIQS-11)?",
        "options_uz": [
            "Faqat uskunaning maksimal quvvati va aylanish tezligini",
            "Xavfsizlik, sifat, operatsiya ketma-ketligi va takt vaqtini",
            "Brigadirning shaxsiy tajribasi va tavsiyalarini",
            "Faqat 5S tozalik qoidalari va asboblar joylashuvini"
        ],
        "options_ru": [
            "Только максимальную мощность оборудования и скорость вращения",
            "Требования безопасности, качества, порядок операций и время такта",
            "Личный опыт и индивидуальные рекомендации бригадира",
            "Только правила уборки 5S и расположение инструментов"
        ],
        "correct": 1,
        "explanation_uz": "BIQS-11: SOS va JES kartalari xavfsizlik, sifat va vaqt bo'yicha aniq ketma-ketlikni belgilaydi.",
        "explanation_ru": "BIQS-11: Карты SOS/JES определяют безопасную, качественную и эффективную последовательность действий."
    },
    {
        "id": 12,
        "question_uz": "4M o'zgarishlar nazorati (4M Control) qaysi omillarni o'z ichiga oladi (BIQS-12)?",
        "question_ru": "Какие 4 фактора входят в управление изменениями 4M (BIQS-12)?",
        "options_uz": [
            "Odam (Man), Mashina (Machine), Material (Material), Jarayon (Method)",
            "Boshqaruv (Manager), Oylik (Money), Bozor (Market), O'lchov (Measurement)",
            "Mashina (Machine), Motor (Motor), Moy (Oil), Usta (Master)",
            "Muddati (Minutes), Mahsulot (Material), Metod (Method), Oylik (Money)"
        ],
        "options_ru": [
            "Человек (Man), Оборудование (Machine), Материал (Material), Метод (Method)",
            "Менеджер (Manager), Деньги (Money), Рынок (Market), Измерение (Measurement)",
            "Машина, Мотор, Масло, Мастер",
            "Минуты, Материал, Метод, Монетизация"
        ],
        "correct": 0,
        "explanation_uz": "BIQS-12: 4M — Odam (Man), Mashina (Machine), Material (Material), Jarayon (Method) o'zgarishlaridir.",
        "explanation_ru": "BIQS-12: 4M включает контроль изменений по Человеку, Оборудованию, Материалу и Процессу."
    },
    {
        "id": 13,
        "question_uz": "Sifat tekshiruv vositalarining yetarliligi va ulardan foydalanish qaysi elementda ko'rib chiqiladi (BIQS-13)?",
        "question_ru": "В каком элементе рассматривается достаточность и эффективность средств контроля (BIQS-13)?",
        "options_uz": [
            "BIQS-07 (O'lchov vositalari poverkasi va MSA)",
            "BIQS-01 (Nomuvofiq mahsulotni ajratish)",
            "BIQS-13 (Vositalar yetarliligi va o'lchash samaradorligi)",
            "BIQS-21 (QCOS Torque va Kritik nuqtalar)"
        ],
        "options_ru": [
            "BIQS-07 (Поверка приборов и анализ MSA)",
            "BIQS-01 (Изоляция брака)",
            "BIQS-13 (Достаточность и эффективность средств контроля)",
            "BIQS-21 (Критические точки QCOS)"
        ],
        "correct": 2,
        "explanation_uz": "BIQS-13: Sifatni tekshirish vositalari (shtangensirkul, shablonlar) yetarli va samarali bo'lishi kerak.",
        "explanation_ru": "BIQS-13: Оценка эффективности средств контроля и умения рабочих ими пользоваться."
    },
    {
        "id": 14,
        "question_uz": "Jarayon yoki detaldagi o'zgarishlarda qaysi blanka to'ldiriladi (BIQS-14)?",
        "question_ru": "Какой бланк заполняется при любых изменениях деталей или процессов на участке (BIQS-14)?",
        "options_uz": [
            "PTR blanqasi to'ldiriladi va Breakpoint nuqtasi belgilanadi",
            "Faqat PFMEA xatarlar jadvalida ball o'zgartiriladi",
            "Bypass Management yo'riqnomasi rasmiylashtiriladi",
            "Faqat Control Plan (CP) hujjati qayta chop etiladi"
        ],
        "options_ru": [
            "Оформляется бланк PTR и фиксируется точка Breakpoint",
            "Изменяется только балл риска в таблице PFMEA",
            "Оформляется временная инструкция Bypass Management",
            "Перепечатывается только План контроля (CP)"
        ],
        "correct": 0,
        "explanation_uz": "BIQS-14: Har bir o'zgarishda PTR to'ldirilib, 4M doskasida Breakpoint qayd etiladi.",
        "explanation_ru": "BIQS-14: Изменения требуют оформления PTR и фиксации точек Breakpoint."
    },
    {
        "id": 15,
        "question_uz": "Ishlab chiqarishda favqulodda muammo aniqlanganda xabardor qilish tizimi (BIQS-15)?",
        "question_ru": "Как называется система оперативного оповещения о проблемах на линии (BIQS-15)?",
        "options_uz": [
            "Fast Response shoshilinch tizimi",
            "ANDON tizimi (Signal/Chaqiruv va ogohlantirish)",
            "Poka-Yoke 3-pog'ona blokirovkasi",
            "Layered Audit (LA) operativ daftari"
        ],
        "options_ru": [
            "Система экстренного реагирования Fast Response",
            "Система ANDON (Свето-звуковой вызов и оповещение)",
            "Автоматическая блокировка Poka-Yoke 3-го уровня",
            "Лист оперативного аудита Layered Audit (LA)"
        ],
        "correct": 1,
        "explanation_uz": "BIQS-15: ANDON tizimi — muammo yuzaga kelganda liniyani to'xtatib ustani chaqirishni ta'minlaydi.",
        "explanation_ru": "BIQS-15: ANDON — это свето-звуковая система вызова мастера при проблемах."
    },
    {
        "id": 16,
        "question_uz": "Muammo limitdan oshganda kimlarga xabar berilishi kerak (BIQS-16)?",
        "question_ru": "Кому сообщается о проблеме при превышении лимитов эскалации (BIQS-16)?",
        "options_uz": [
            "Faqat mehnat muhofazasi va xavfsizlik inspektoriga",
            "Eskalatsiya reglamenti bo'yicha tegarli rahbarlar va mutaxassislarga",
            "Tashqi ta'minotchilar (SUB) va IQC bo'limiga",
            "Faqat sifat nazoratchisi (QC) va smena operatoriga"
        ],
        "options_ru": [
            "Только инспектору по охране труда и технике безопасности",
            "Соответствующим руководителям и специалистам по матрице эскалации",
            "Внешним поставщикам (SUB) и отделу IQC",
            "Только контролеру качества (QC) и оператору смены"
        ],
        "correct": 1,
        "explanation_uz": "BIQS-16: Muammo mezonidan oshsa, belgilangan tartibda rahbarlarga (eskalatsiya) xabar qilinadi.",
        "explanation_ru": "BIQS-16: Регламент эскалации требует вызова руководителей при превышении лимитов брака."
    },
    {
        "id": 17,
        "question_uz": "Vizual boshqaruv (Visual Management) standarti nima uchun kerak (BIQS-17)?",
        "question_ru": "Для чего нужны стандарты визуального менеджмента (BIQS-17)?",
        "options_uz": [
            "Faqat ISO 9001 auditorlariga namoyish etish uchun",
            "OK va NG holatlarini tez farqlash va chetlanishlarni ko'rish uchun",
            "Ishchilarning nobay soatlik vaqtini hisoblash uchun",
            "Sexdagi asboblarning bozor narxini ko'rsatish uchun"
        ],
        "options_ru": [
            "Исключительно для демонстрации аудиторам ISO 9001",
            "Для оперативного визуального отличия нормы (OK) от брака (NG)",
            "Для учета невыходов и рабочего времени сотрудников",
            "Для отображения стоимости оборудования цеха"
        ],
        "correct": 1,
        "explanation_uz": "BIQS-17: Vizual boshqaruv orqali ish joyidagi har qanday og'ish va nosozlik bir qarashda ko'rinadi.",
        "explanation_ru": "BIQS-17: Визуальный менеджмент позволяет с первого взгляда отличить норму (OK) от отклонения (NG)."
    },
    {
        "id": 18,
        "question_uz": "Sifat ko'rgazmali qo'llanmalari va o'zgarishlar qanday yetkaziladi (BIQS-18)?",
        "question_ru": "Как должны доводиться до сотрудников визуальные инструкции и изменения (BIQS-18)?",
        "options_uz": [
            "Ish joyiga ko'rinarli o'rnatiladi va xodim amaliy o'qitiladi",
            "Faqat tsex boshlig'i xonasidagi arxiv papkasida saqlanadi",
            "Smena boshida faqat og'zaki tarzda o'qib eshitiriladi",
            "Zavodning ichki korporativ veb-saytiga joylashtiriladi"
        ],
        "options_ru": [
            "Размещаются непосредственно на рабочем месте с практическим обучением",
            "Хранятся исключительно в архивной папке начальника цеха",
            "Зачитываются исключительно устно в начале смены",
            "Размещаются только на внутреннем портале завода"
        ],
        "correct": 0,
        "explanation_uz": "BIQS-18: Barcha yo'riqnomalar ish joyida vizual ko'rinishda osilgan va xodim o'qitilgan bo'lishi shart.",
        "explanation_ru": "BIQS-18: Инструкции должны висеть прямо перед глазами оператора."
    },
    {
        "id": 19,
        "question_uz": "Jarayon nazorati uchun qaysi hujjatlar bir-biriga mos bo'lishi shart (BIQS-19)?",
        "question_ru": "Какие документы контроля процесса должны полностью соответствовать друг другу (BIQS-19)?",
        "options_uz": [
            "PPR grafiki, TPM checklisti, FIFO va SCOS",
            "Control Plan (CP), Flowchart, FMEA va SOS/JES",
            "PTR blanqasi, Breakpoint, Bypass va Andon",
            "MSA tahlili, IQC kirish nazorati va LA varog'i"
        ],
        "options_ru": [
            "График ППР, чек-лист TPM, FIFO и SCOS",
            "План контроля (CP), Flowchart, FMEA и SOS/JES",
            "Бланк PTR, Breakpoint, Bypass и Andon",
            "Анализ MSA, входной контроль IQC и лист LA"
        ],
        "correct": 1,
        "explanation_uz": "BIQS-19: Texnologik jarayondagi barcha asosiy sifat hujjatlari (CP, FMEA, SOS) bir-biriga 100% mos kelishi kerak.",
        "explanation_ru": "BIQS-19: CP, FMEA и SOS — это связка документов, которые должны быть синхронизированы."
    },
    {
        "id": 20,
        "question_uz": "Ish joyida xodimning amaliy harakatlari qaysi hujjatga mosligini tekshirish kerak (BIQS-20)?",
        "question_ru": "Соответствие каким документам нужно проверять при оценке работы оператора на линии (BIQS-20)?",
        "options_uz": [
            "Ishchining mehnat shartnomasi va lavozim yo'riqnomasiga mosligi",
            "Amaliy ishning SOS/JES, CheckList va Control Plan talablariga mosligi",
            "Uskunaning texnik pasporti va kalibrovka guvohnomasi",
            "Ta'minotchining IQC kirish hujjatlari va hisob-fakturasi"
        ],
        "options_ru": [
            "Соответствие трудового договора и должностной инструкции",
            "Строгое соответствие реальных действий картам SOS/JES, чек-листам и План-контролю",
            "Технический паспорт и свидетельство поверки станка",
            "Входная накладная IQC и счет-фактура поставщика"
        ],
        "correct": 1,
        "explanation_uz": "BIQS-20: Operator ishni faqat SOS va CheckList hujjatlariga asosan xatosiz bajarishi shart.",
        "explanation_ru": "BIQS-20: Выполнение операций проверяется на строгое соответствие SOS и чек-листам."
    },
    {
        "id": 21,
        "question_uz": "Kritik va xavfli texnologik nuqtalar (QCOS Torque/Weld) qanday nazorat qilinadi (BIQS-21)?",
        "question_ru": "Как контролируются критические и опасные точки техпроцесса (QCOS Torque/Weld) (BIQS-21)?",
        "options_uz": [
            "Vaqtida o'lchanadi va SPC (Statistik jarayon nazorati) o'tkaziladi",
            "Faqat mahsulot mijozga yuborilgandan so'ng sinalladi",
            "Yilda bir marta ISO audit vaqtida tekshiriladi",
            "Faqat vizual ko'zdan kechirilib, markerovka qo'yiladi"
        ],
        "options_ru": [
            "Регулярно замеряются с проведением статистического анализа SPC",
            "Проверяются только перед отправкой готовой продукции клиенту",
            "Проверяются раз в год во время аудита ISO",
            "Оцениваются исключительно визуально с нанесением метки"
        ],
        "correct": 0,
        "explanation_uz": "BIQS-21: Payvand choki, qotirish momenti kabi kritik nuqtalar maxsus SPC orqali kuzatib boriladi.",
        "explanation_ru": "BIQS-21: Моменты затяжки и точки сварки критически важны для безопасности."
    },
    {
        "id": 22,
        "question_uz": "Ta'mirlash (Rework) operatsiyalari qay tartibda amalga oshiriladi (BIQS-22)?",
        "question_ru": "В каком порядке выполняются операции доработки/ремонта деталей (Rework) (BIQS-22)?",
        "options_uz": [
            "Konveyer harakatlanayotgan vaqtda bevosita ish joyida",
            "Alohida zonada, maxsus SOS va Flexibility chart bo'yicha malakali xodim tomonidan",
            "Smena oxirida xohlagan operator tomonidan",
            "Ta'minotchi (SUB) vakili kelishini kutib omborda"
        ],
        "options_ru": [
            "Прямо на движущемся конвейере на рабочем месте",
            "В изолированной зоне обученным персоналом по SOS и матрице Flexibility",
            "Любым свободным оператором в конце смены",
            "На складе в ожидании представителя поставщика (SUB)"
        ],
        "correct": 1,
        "explanation_uz": "BIQS-22: Rework (qayta ishlash) faqat ruxsat etilgan xodim tomonidan va alohida sektorda bajariladi.",
        "explanation_ru": "BIQS-22: Доработка деталей выполняется только обученным персоналом в специальной зоне."
    },
    {
        "id": 23,
        "question_uz": "Sifat muammolari haqida boshqa tsexlarga qanday xabar beriladi (BIQS-23)?",
        "question_ru": "Как смежные участки оповещаются о проблемах качества (BIQS-23)?",
        "options_uz": [
            "Faqat oylik Fast Response hisobotida ko'rsatish orqali",
            "Quality Alert va Containment Sheet orqali oldinga va orqaga zudlik bilan",
            "Faqat telefon orqali og'zaki xabar berish bilan",
            "LPA auditi o'tkazish vaqtida daftarga yozish orqali"
        ],
        "options_ru": [
            "Исключительно в ежемесячном отчете Fast Response",
            "Двусторонним оперативным уведомлением Quality Alert и Containment Sheet",
            "Только устным звонком по телефону",
            "Путем записи в журнал во время проведения аудита LPA"
        ],
        "correct": 1,
        "explanation_uz": "BIQS-23: Nuqson topilganda uni manbayi va qabul qiluvchisi Quality Alert (Containment) bilan ogohlantiriladi.",
        "explanation_ru": "BIQS-23: Quality Alert гарантирует, что брак не уйдет к клиенту и поставщик узнает о дефекте."
    },
    {
        "id": 24,
        "question_uz": "Xodimlarning operatsiyalarni bajarish malakasi (Flexibility chart) qanday ta'minlanadi (BIQS-24)?",
        "question_ru": "Как обеспечивается и контролируется квалификация рабочих (Flexibility chart) (BIQS-24)?",
        "options_uz": [
            "Oliy ma'lumot diplomi va umumiy mehnat stajiga qarab",
            "SOS/JES asosida JIT o'qitish va Flexibility chart orqali tasdiqlash",
            "Operatorning shaxsiy xohishi va o'z-o'zini baholashi orqali",
            "Mehnat muhofazasi bo'limining umumiy kirish blanqasi bo'yicha"
        ],
        "options_ru": [
            "На основе диплома о высшем образовании и общего стажа",
            "Обучение JIT по SOS/JES и подтверждение в матрице Flexibility chart",
            "По личному заявлению и самооценке оператора",
            "По общему вводному инструктажу отдела охраны труда"
        ],
        "correct": 1,
        "explanation_uz": "BIQS-24: Har bir xodim faqat o'ziga o'rgatilgan (JIT) va matritsada tasdiqlangan ishni bajara oladi.",
        "explanation_ru": "BIQS-24: Матрица навыков подтверждает, что рабочий обучен стандарту."
    },
    {
        "id": 25,
        "question_uz": "Ish joyida mahsulotni chang va kirdan himoya qilish tartibi qaysi elementga kiradi (BIQS-25)?",
        "question_ru": "К какому элементу относится защита продукции и рабочего места от пыли и загрязнений (BIQS-25)?",
        "options_uz": [
            "BIQS-25 (Ish joyi va mahsulotni chang hamda ifloslanishdan himoya qilish)",
            "BIQS-17 (Vizual boshqaruv va 5S ko'rgazmasi)",
            "BIQS-26 (TPM uskunalar va PPR texnik xizmati)",
            "BIQS-01 (Nomuvofiq mahsulot red tag zonasi)"
        ],
        "options_ru": [
            "BIQS-25 (Защита рабочего места и продукции от пыли и загрязнений)",
            "BIQS-17 (Визуальный менеджмент и 5S)",
            "BIQS-26 (Техническое обслуживание оборудования TPM и ППР)",
            "BIQS-01 (Красная зона несоответствующей продукции)"
        ],
        "correct": 0,
        "explanation_uz": "BIQS-25: Ish joyida detalga chang tushmasligi uchun tozalik standarti qat'iy ta'minlanishi zarur.",
        "explanation_ru": "BIQS-25: Поддержание идеальной чистоты для предотвращения дефектов внешнего вида."
    },
    {
        "id": 26,
        "question_uz": "Uskunalar sifatli ishlashi uchun nima profilaktika qilinadi (BIQS-26)?",
        "question_ru": "Какая профилактика применяется для безотказной работы оборудования (BIQS-26)?",
        "options_uz": [
            "Uskuna jiddiy buzilib, konveyer to'xtaganda",
            "TPM checklistlari va PPR rejali ta'mirlash grafikalari asosida",
            "Faqat Tashqi kalibrovka tashkiloti auditi vaqtida",
            "Smena boshlig'ining kayfiyati va topshirig'iga qarab"
        ],
        "options_ru": [
            "Исключительно после серьезной поломки и остановки линии",
            "По чек-листам TPM и графикам планово-предупредительного ремонта (ППР)",
            "Только во время аудита внешней калибровочной организации",
            "По индивидуальному указанию начальника смены"
        ],
        "correct": 1,
        "explanation_uz": "BIQS-26: TPM operator tomonidan, PPR mexaniklar tomonidan vaqtida bajarilishi uskunani asraydi.",
        "explanation_ru": "BIQS-26: Регулярное обслуживание TPM/ППР предотвращает внезапные простои и брак."
    },
    {
        "id": 27,
        "question_uz": "FIFO qoidasi nima uchun muhim (BIQS-27)?",
        "question_ru": "Для чего критически важно соблюдение правила FIFO (BIQS-27)?",
        "options_uz": [
            "Konteynerlarni sexda tekshirish vaqtini tejash uchun",
            "Birinchi kelgan detal birinchi ishlatilib, eskirib va sifati buzilib qolmasligi uchun",
            "Ombordagi bo'sh maydonni maksimal kamaytirish uchun",
            "Faqat detallarning rangi va shakli bo'yicha guruhlash uchun"
        ],
        "options_ru": [
            "Для сокращения времени проверки контейнеров в цехе",
            "Первая поступившая деталь используется первой для предотвращения порчи и старения",
            "Для максимального уменьшения свободной площади склада",
            "Исключительно для группировки деталей по цвету и форме"
        ],
        "correct": 1,
        "explanation_uz": "BIQS-27: FIFO (First-In, First-Out) zaxiralar yaroqlilik muddatini nazorat qilish kafolatidir.",
        "explanation_ru": "BIQS-27: FIFO предотвращает старение и порчу компонентов на складе."
    },
    {
        "id": 28,
        "question_uz": "Mahsulotlar qanday idish (tara) larda yetkazilishi shart (BIQS-28)?",
        "question_ru": "В какой таре должна поставляться и храниться продукция (BIQS-28)?",
        "options_uz": [
            "Sexda topilgan va bo'shagan har qanday karton va yog'och qutilarda",
            "Limitsiz miqdorda to'g'ridan-to mezonda polda to'plangan holda",
            "Tasdiqlangan maxsus konteyner, yorliq (birka) va Min/Max limitlari asosida",
            "Faqat etkazib beruvchining ochiq plastik paketlarida"
        ],
        "options_ru": [
            "В любой освободившейся картонной или деревянной таре",
            "Навалом прямо на полу без ограничения количества",
            "В утвержденной спец-таре с бирками и соблюдением лимитов Min/Max",
            "Только в открытых пластиковых пакетах поставщика"
        ],
        "correct": 2,
        "explanation_uz": "BIQS-28: Noto'g'ri tara mahsulot sifatiga zarar yetkazadi, shuning uchun faqat tasdiqlangan taralar ruxsat etiladi.",
        "explanation_ru": "BIQS-28: Использование нестандартной тары ведет к повреждению деталей."
    },
    {
        "id": 29,
        "question_uz": "Ta'minotchilardan (SUB) kelayotgan ehtiyot qismlar qanday nazorat qilinadi (BIQS-29)?",
        "question_ru": "Как контролируются компоненты, поступающие от субпоставщиков (BIQS-29)?",
        "options_uz": [
            "Faqat detallarning bojxona deklaratsiyasi va yuk xati asosida",
            "IQC kirish nazorati, sifat yorliqlari va BIQS 1-13 auditlari orqali",
            "Ta'minotchining o'zi tomonidan berilgan sifat kafolat xati bo'yicha",
            "Tayyor avtomobilni oxirgi sinov trekida haydab ko'rish orqali"
        ],
        "options_ru": [
            "Только на основании таможенной декларации и накладной",
            "Через входной контроль IQC, маркировку качества и аудиты BIQS 1-13",
            "По гарантийному письму самого поставщика без проверки",
            "При контрольном испытании готового автомобиля на треке"
        ],
        "correct": 1,
        "explanation_uz": "BIQS-29: Kirib kelayotgan qismlar sifatli bo'lmas ekan, yakuniy mahsulot ham sifatli bo'lmaydi.",
        "explanation_ru": "BIQS-29: Строгий входной контроль IQC отсеивает брак от поставщиков."
    },
    {
        "id": 30,
        "question_uz": "Mehnat xavfsizligi va GMS PI qoidalari qaysi BIQS standartida yozilgan (BIQS-30)?",
        "question_ru": "В каком стандарте BIQS описаны требования безопасности и GMS PI (BIQS-30)?",
        "options_uz": [
            "BIQS-01 (Nomuvofiq mahsulot red tag)",
            "BIQS-15 (ANDON chaqiruv tizimi)",
            "BIQS-30 (Mehnat muhofazasi, SCOS ko'rgazmali yo'riqnomalari va GMS PI)",
            "BIQS-26 (TPM uskunalar va PPR)"
        ],
        "options_ru": [
            "BIQS-01 (Изоляция брака)",
            "BIQS-15 (Система Andon)",
            "BIQS-30 (Охрана труда, наглядные инструкции SCOS и требования GMS PI)",
            "BIQS-26 (TPM и ППР)"
        ],
        "correct": 2,
        "explanation_uz": "BIQS-30: Xavfsizlik har doim birinchi o'rinda! SCOS qoidalariga hamma amal qilishi shart.",
        "explanation_ru": "BIQS-30: Безопасность на первом месте. Выполнение инструкций SCOS обязательно для всех."
    }
]

if __name__ == "__main__":
    init_db()
    print("Database initialized successfully.")
